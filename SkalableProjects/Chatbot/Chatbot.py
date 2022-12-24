import openpyxl

while True:
    val = input("Ask your question: ").lower()
    #print(val)

    theFile = openpyxl.load_workbook('Questions.xlsx')
    allSheetNames = theFile.sheetnames

    #print("All sheet names {} " .format(theFile.sheetnames))

    found = False

    for sheet in allSheetNames:
        #print("Current sheet name is {}" .format(sheet))
        currentSheet = theFile[sheet]

        for row in range(1, currentSheet.max_row + 1):
            #print(row)
            for column in "A":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                #print("val is " + val)
                #print("cell name is " + currentSheet[cell_name].value)
                if (currentSheet[cell_name].value in val):
                    found = True
                    print(currentSheet["{}{}".format("B", row)].value)